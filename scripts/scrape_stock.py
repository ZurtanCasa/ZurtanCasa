#!/usr/bin/env python3
"""
Zeta Software — scraper de stock actual (muebles + alfombras neutras).
Requiere: ZETA_USER, ZETA_PASS

URL confirmada: z.gestion.reportes.stockactual
Campos del formulario:
  vLOCIDART  → Local empresa  (1=Casa Central, 2=Dialcaren)
  vDEPIDSA   → Depósito       (1=Local, 2=Dialcaren, 3=Milanco, 4=Reserva)
  vCATEGARTID→ Categoría      (vacío=todas)
  vGENERARXLS→ checkbox XLS
  BTNENTER   → Confirmar
"""
import os, json, sys, tempfile, shutil, time, re
from datetime import datetime, timezone, timedelta

DATA_PATH = os.path.join(os.path.dirname(__file__), "..", "data", "stock.json")
BASE_URL  = "https://www.zetasoftware.com/z.info.inicio"
STOCK_URL = "https://www.zetasoftware.com/z.gestion.reportes.stockactual"
UY_TZ     = timezone(timedelta(hours=-3))

# Categorías de neutras según Zeta
NEUTRAS_CATS = {"pura lana", "yute y lana", "yute y lana diseño", "yute y lana nuevas", "exterior pet"}

def uy_now():
    return datetime.now(UY_TZ)

# ─── Login ────────────────────────────────────────────────────────────────────

def login(page, user, password):
    # Login robusto (mismo flujo que scrape_articulos.py, ya probado):
    # maneja tanto el form email+pass en una pantalla como el flujo email-first.
    page.goto(BASE_URL, timeout=60000, wait_until="domcontentloaded")
    try:
        page.wait_for_load_state("networkidle", timeout=30000)
    except Exception:
        pass

    print(f"  [login] URL: {page.url}")

    # ¿El campo de password ya está en el DOM?
    has_pw = page.evaluate(
        "() => !!document.querySelector(\"input[name='vSIMPLEPASSWORD']\")"
    )
    if not has_pw:
        # Flujo email-first: el password aparece solo tras confirmar el email
        print("  [login] Password no visible, probando flujo email-first...")
        email_el = page.wait_for_selector("input[name='vUSULIBRAEMAIL']", timeout=30000)
        email_el.fill(user)
        email_el.press("Enter")
        time.sleep(4)

    page.wait_for_selector("input[name='vSIMPLEPASSWORD']", timeout=60000)
    page.fill("input[name='vUSULIBRAEMAIL']", user, timeout=30000)
    page.fill("input[name='vSIMPLEPASSWORD']", password, timeout=30000)
    page.click("input[name='BTNENTER']")
    time.sleep(10)
    frame = page.frames[0]
    print(f"  Login OK — {frame.url}")
    return frame

# ─── Download helpers ─────────────────────────────────────────────────────────

def _row0001_ready(frame) -> bool:
    """True si la fila 0001 (nuestro informe recién generado) ya ofrece 'Descargar XLS'."""
    return bool(frame.evaluate("""() => {
        const sel = document.querySelector('select[name="vACCIONES_0001"]');
        if (!sel) return false;
        return Array.from(sel.options).some(o => o.text.includes('XLS'));
    }"""))


def wait_and_download(frame, page, dest: str, timeout_sec=420) -> bool:
    # Nuestro informe recién generado es SIEMPRE la fila más nueva = vACCIONES_0001.
    # Mientras procesa, esa fila solo ofrece "Eliminar"; al terminar aparece
    # "Descargar XLS". Recargamos procesosww hasta que 0001 esté listo.
    print(f"    Esperando que el informe (fila 0001) termine (máx {timeout_sec}s)...")
    deadline = time.time() + timeout_sec
    ready = False
    while time.time() < deadline:
        try:
            if _row0001_ready(frame):
                ready = True
                break
        except Exception:
            pass
        time.sleep(5)
        # Recargar para que procesosww actualice el estado del informe
        try:
            frame.goto("https://www.zetasoftware.com/z.informes.procesosww",
                       wait_until="domcontentloaded", timeout=30000)
            time.sleep(2)
        except Exception:
            pass

    if not ready:
        # Diagnóstico final
        estado = frame.evaluate("""() => {
            const sel = document.querySelector('select[name="vACCIONES_0001"]');
            return sel ? Array.from(sel.options).map(o => o.text) : 'sin fila 0001';
        }""")
        print(f"    ⚠ La fila 0001 no llegó a ofrecer XLS a tiempo. Opciones: {estado}")
        return False

    print(f"    ✅ Informe terminado — descargando fila 0001...")

    # Descargar la fila 0001 vía execEvt (patrón probado en scrape_articulos)
    gxoch = frame.evaluate("""() => {
        const sel = document.querySelector('select[name="vACCIONES_0001"]');
        if (!sel) return null;
        const opt = Array.from(sel.options).find(o => o.text.includes('XLS'));
        return {
            name:   'vACCIONES_0001',
            value:  opt ? opt.value : '2',
            gxoch0: sel.getAttribute('data-gxoch0'),
        };
    }""")
    print(f"    gxoch={gxoch}")

    if gxoch and gxoch.get("gxoch0"):
        m = re.search(r"execEvt\('([^']*)',([^,]*),\s*'([^']*)',this,(\d+)\)", gxoch["gxoch0"])
        if m:
            p1, p2, evtname, rowid = m.group(1), m.group(2), m.group(3), m.group(4)
            xls_val = gxoch["value"]
            print(f"    execEvt params: evtname={evtname!r} rowid={rowid} val={xls_val}")
            try:
                with page.expect_download(timeout=120000) as dl_info:
                    frame.evaluate(f"""() => {{
                        const sel = document.querySelector('select[name="{gxoch['name']}"]');
                        if (!sel) return;
                        sel.value = '{xls_val}';
                        gx.evt.execEvt('{p1}', {p2}, '{evtname}', sel, {rowid});
                    }}""")
                dl_info.value.save_as(dest)
                print(f"    ✅ Descargado (execEvt) {gxoch['name']} ({os.path.getsize(dest):,} bytes)")
                return True
            except Exception as e:
                print(f"    ⚠ execEvt/download falló: {e}")
        else:
            print(f"    ⚠ No se pudo parsear gxoch0: {gxoch['gxoch0']!r}")

    print(f"    ⚠ Todas las estrategias de descarga fallaron")
    return False

# ─── Fetch por depósito ───────────────────────────────────────────────────────

def fetch_for_deposito(frame, page, deposito_text: str, tmp_dir: str, fname: str) -> str | None:
    """Genera y descarga el Excel de stock para un depósito. Devuelve ruta o None."""
    dest = os.path.join(tmp_dir, fname)

    print(f"  Navegando a Stock Actual ({deposito_text})...")
    frame.goto(STOCK_URL, wait_until="domcontentloaded", timeout=30000)
    try:
        frame.wait_for_load_state("networkidle", timeout=15000)
    except Exception:
        pass
    time.sleep(2)

    # Configurar formulario
    result = frame.evaluate(f"""() => {{
        function setSelect(name, text) {{
            const el = document.querySelector('[name="' + name + '"]');
            if (!el || el.tagName !== 'SELECT') return 'not found: ' + name;
            const opt = Array.from(el.options).find(o =>
                o.text.toLowerCase().includes(text.toLowerCase())
            );
            if (!opt) return 'option not found: ' + text + ' in ' + name;
            el.value = opt.value;
            el.dispatchEvent(new Event('change', {{bubbles: true}}));
            el.dispatchEvent(new Event('blur',   {{bubbles: true}}));
            return name + '=' + opt.value + ':' + opt.text.trim();
        }}
        function clearSelect(name) {{
            const el = document.querySelector('[name="' + name + '"]');
            if (!el || el.tagName !== 'SELECT') return;
            el.value = '';
            el.dispatchEvent(new Event('change', {{bubbles: true}}));
        }}
        // Local siempre = Casa Central
        const r1 = setSelect('vLOCIDART', 'Casa Central');
        // Depósito según parámetro
        const r2 = setSelect('vDEPIDSA', '{deposito_text}');
        // Limpiar filtro de categoría para obtener todas
        clearSelect('vCATEGARTID');
        // Activar Generar XLS
        const xls = document.querySelector('[name="vGENERARXLS"]');
        if (xls && xls.type === 'checkbox' && !xls.checked)
            xls.dispatchEvent(new MouseEvent('click', {{bubbles: true, cancelable: true}}));
        return {{local: r1, deposito: r2}};
    }}""")
    print(f"  Formulario: {result}")

    if result and "not found" in str(result.get("deposito", "")):
        print(f"  ⚠ Depósito '{deposito_text}' no encontrado — abortando")
        return None

    time.sleep(1)

    # Submit
    has_enter = frame.evaluate("() => !!document.querySelector('[name=\"BTNENTER\"]')")
    if has_enter:
        try:
            frame.click("[name='BTNENTER']", timeout=15000)
            page.wait_for_url("**/z.informes.procesosww**", timeout=120000)
            print("  ✅ En procesosww")
            ok = wait_and_download(frame, page, dest)
            return dest if ok else None
        except Exception as e:
            print(f"  ⚠ BTNENTER falló: {e}")

    # Fallback BTNEXPORT
    has_export = frame.evaluate("() => !!document.querySelector('input#BTNEXPORT, [name=\"BTNEXPORT\"]')")
    if has_export:
        try:
            with page.expect_download(timeout=60000) as dl_info:
                frame.click("input#BTNEXPORT, [name='BTNEXPORT']", timeout=10000)
            dl_info.value.save_as(dest)
            return dest
        except Exception as e:
            print(f"  ⚠ BTNEXPORT falló: {e}")

    print("  ⚠ Sin botón de submit")
    return None

# ─── Parseo de Excel ──────────────────────────────────────────────────────────

def parse_stock_excel(path: str) -> dict[str, int]:
    """
    Parsea Excel de Stock Actual.
    Filtra muebles (cat='Muebles') y neutras (cat en NEUTRAS_CATS).
    Devuelve {codigo: cantidad}.
    """
    import openpyxl
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active

    # Detectar fila de encabezados buscando "ARTICULO"
    header_row = 5
    for r in range(1, 10):
        if any("ARTICULO" in str(ws.cell(r, c).value or "") for c in range(1, 5)):
            header_row = r
            break

    # Mapear columnas desde encabezados
    headers = {str(ws.cell(header_row, c).value or "").strip().upper(): c
               for c in range(1, ws.max_column + 1)}
    print(f"    Encabezados: {dict(list(headers.items())[:12])}")

    col_cod  = headers.get("ARTICULO", 1)
    col_nom  = headers.get("NOMBRE",   2)
    col_cat  = headers.get("CATEGORÍA", headers.get("CATEGORIA", 6))
    # Stock: última columna numérica con datos (generalmente la última)
    col_stk  = ws.max_column

    result: dict[str, int] = {}
    skipped = 0
    for r in range(header_row + 1, ws.max_row + 1):
        codigo   = ws.cell(r, col_cod).value
        categoria = str(ws.cell(r, col_cat).value or "").strip().lower()
        stock_val = ws.cell(r, col_stk).value

        if not codigo:
            continue

        is_mueble = "mueble" in categoria
        is_neutra = categoria in NEUTRAS_CATS

        if not (is_mueble or is_neutra):
            skipped += 1
            continue

        try:
            qty = int(float(stock_val)) if stock_val is not None else 0
        except (ValueError, TypeError):
            qty = 0

        result[str(codigo).strip()] = qty

    print(f"    Parseados: {len(result)} artículos (muebles+neutras), {skipped} omitidos")
    return result

# ─── Main ─────────────────────────────────────────────────────────────────────

def main():
    user     = os.environ.get("ZETA_USER", "")
    password = os.environ.get("ZETA_PASS", "")

    if not user or not password:
        print("SKIP: ZETA_USER y ZETA_PASS requeridos", file=sys.stderr)
        sys.exit(0)

    try:
        from playwright.sync_api import sync_playwright
        import openpyxl  # noqa: F401
    except ImportError as e:
        print(f"ERROR: {e}", file=sys.stderr)
        sys.exit(1)

    tmp_dir = tempfile.mkdtemp()

    try:
        with sync_playwright() as p:
            browser = p.chromium.launch(headless=True)
            page    = browser.new_page()
            frame   = login(page, user, password)

            print("\n[1/2] Stock Local (Casa Central)...")
            path_local = fetch_for_deposito(frame, page,
                deposito_text="Local",
                tmp_dir=tmp_dir,
                fname="stock_local.xlsx",
            )

            print("\n[2/2] Stock Dialcaren...")
            path_dialcaren = fetch_for_deposito(frame, page,
                deposito_text="Dialcaren",
                tmp_dir=tmp_dir,
                fname="stock_dialcaren.xlsx",
            )

            browser.close()

        local_data     = parse_stock_excel(path_local)     if path_local     else {}
        dialcaren_data = parse_stock_excel(path_dialcaren) if path_dialcaren else {}

        all_codes = set(local_data) | set(dialcaren_data)
        articulos = {
            cod: {
                "local":     local_data.get(cod, 0),
                "dialcaren": dialcaren_data.get(cod, 0),
            }
            for cod in all_codes
        }

        output = {
            "_status":               "ok",
            "_ultima_actualizacion": uy_now().strftime("%Y-%m-%dT%H:%M:%S"),
            "_tiene_datos":          bool(articulos),
            "articulos":             articulos,
        }

        with open(DATA_PATH, "w", encoding="utf-8") as f:
            json.dump(output, f, ensure_ascii=False, indent=2)

        total_stock = sum(v["local"] + v["dialcaren"] for v in articulos.values())
        print(f"\n✅ stock.json: {len(articulos)} artículos, {total_stock} unidades totales")

    except Exception as e:
        print(f"ERROR: {e}", file=sys.stderr)
        import traceback; traceback.print_exc()
        if not os.path.exists(DATA_PATH):
            with open(DATA_PATH, "w") as f:
                json.dump({"_status": "error", "_tiene_datos": False, "articulos": {}}, f)
        sys.exit(1)
    finally:
        shutil.rmtree(tmp_dir, ignore_errors=True)

if __name__ == "__main__":
    main()
